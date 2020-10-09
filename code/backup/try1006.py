# author:wzt
import pandas as pd

import openpyxl
from numpy import *


# f = "0924wj_2020_09_24_200721_002_Odau_1.xlsx"
# print(f[-15:-12])

wb = openpyxl.Workbook()
ws = wb.active
a = [1,2,3]
ws.append(a)



a = range(7)
b = (3,4,5)
print(type(a))
print(a[2])
print(b)
wb.save("charttest.xlsx")


#
#
# from openpyxl.chart import (
#     Reference,
#     Series,
#     PieChart,
#     BarChart,
#     BubbleChart,
#     LineChart
# )
#
# # Pie chart
#
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "PieChart"
# wb2 = openpyxl.Workbook()
# ws2 = wb2.active
#
# data = [
#     ["pie", "sold"],
#     ["Apple", 50],
#     ["Cherry", 30],
#     ["Pumpkin", 10],
#     ["Chocolate", 40]
# ]
#
# for row in data:
#     ws.append(row)
#
# pie = PieChart()
# labels = Reference(ws, min_col=1, min_row=2)
# data = Reference(ws, min_col=2, min_row=2)
# pie.add_data(data)
# pie.set_categories(labels)
# pie.title = "Pie sold by category"
# ws.add_chart(pie, "A15")
#
# ws = wb.create_sheet("barchart")
# rows = [
#     ("Num", "Batch_1", "batch_2"),
#     (1, 2, 3),
#     (2, 3, 4),
#     (3, 4, 5),
#     (4, 5, 6)
# ]
#
#
#
# for row in rows:
#     ws.append(row)
# chart1 = LineChart()
# chart1.type = "col"
# chart1.style = 15
# chart1.title = "Bar"
# chart1.y_axis.title = "batch"
# cats = Reference(ws, min_col=1, min_row=2, max_row=5)
# data1 = Reference(ws, min_col=2, min_row=2, max_row=5)
# data2 = Reference(ws, min_col=3, min_row=2, max_row=5)
# chart1.add_data(data1)
# chart1.add_data(data2)
#
# chart1.set_categories(cats)
# ws.add_chart(chart1, "B5")
# ws2.add_chart(chart1)
#
# wb.save("charttest.xlsx")
# wb2.save("charttest2.xlsx")

# books = pd.read_excel("test1.xlsx",skiprows=9, usecols="D:E",dtype={"ID":str})
# print(books)
# for i in books.index:
#     books["ID"].at[i] = i+1
#
# print(books)

# print(f[26:29])
# file_name = pd.read_excel("邬如靖.xlsx", dtype=str)
# print(file_name.columns[2])
# print(file_name)
#
# print(file_name.iloc[1, 3])
#
# src_file_name = f[26:29]
#
# src_file_name_list = []
# for motion in range(0, 6):
#     for motion_pattern in range(2, 5):
#         src_file_name = f[0:26]+file_name.iloc[motion,motion_pattern]+"_Odau_1.xlsx"
#         src_file_name_list.append(src_file_name)
#         new_file_name = f[0:26] + file_name.iloc[motion, motion_pattern] + "_Odau_1" + \
#                                 file_name.columns[motion_pattern] + file_name.iloc[motion, 1] + ".xlsx"
#         print(new_file_name)
# print(src_file_name_list)

#
# for motion in range(1, 7):
#     for type in range(2, 5):
#         new_file_name = f+file_name.columns[type]+file_name.iloc[motion,1]
#
# print(new_file_name)
