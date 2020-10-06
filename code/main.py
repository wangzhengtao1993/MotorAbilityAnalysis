# author:wzt
import pandas as pd
import os
import win32com.client as win32
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
from numpy import *
import openpyxl
from openpyxl.chart import (
    Reference,
    series,
    BarChart,
    LineChart
)


# win32api找不到时 pip install pywin32==225

class EMGProcess(object):

    def __init__(self, folder, subject):
        self.upper_limb = False
        self.header = 4  # 表头在第四行
        self.folder = folder  # 文件夹路径
        # 新建文件夹
        if os.path.exists(self.folder + "\\Process\\"):
            pass
        else:
            os.mkdir(self.folder + "\\Process\\")
        self.process_folder = self.folder + "\\Process\\"
        self.win_set = (100, 100)  # 滑动时间窗
        self.subject = subject
        print("Path:", folder)

    def rename_test_file(self):
        file_list = os.listdir(self.process_folder)
        f = file_list[2]
        print(f[0:26])
        file_name = pd.read_excel(self.subject, dtype=str)
        print(file_name)

        # 遍历所有文件
        if self.upper_limb:
            motion_id_start = 0
            motion_id_end = 6
        else:
            motion_id_start = 7
            motion_id_end = 13

        keyword = "Odau"
        for file in file_list:
            if keyword in file:
                print("file", file[26:29])
                file_id = file[26:29]
                # 遍历所有文件id
                for motion in range(motion_id_start, motion_id_end):
                    for motion_pattern in range(1, 4):
                        # print("motion:", motion)
                        # print("motion_pattern:", motion_pattern)
                        motion_id = file_name.iloc[motion, motion_pattern]
                        # print("motion_id:", motion_id)
                        if file_id == motion_id:
                            dst_name = f[0:26] + file_name.iloc[motion, motion_pattern] + "_Odau_1" + \
                                       file_name.columns[motion_pattern] + file_name.iloc[motion, 0] + ".xlsx"

                            if os.path.exists(self.process_folder + file):
                                os.rename(self.process_folder + file, self.process_folder + dst_name)
                                print("rename:", dst_name)
                            else:
                                pass

        # src_file_name = f[26:29]

    def save_as_high_ver(self):
        """
        源文件格式错误，另存为xlsx
        """
        file_list = os.listdir(self.process_folder)
        if len(file_list) > 0:
            print("文件已转换")
        else:
            # 1.获得当前目录下所有文件名
            file_list = os.listdir(self.folder)
            # 2.打开excel处理程序，固定写法
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            # excel.Application.Visible = False
            print("saving as .xlsx")
            for file in file_list:
                # 将文件名与后缀分开
                file_name, suff = os.path.splitext(file)
                if suff == ".xls":
                    wb = excel.Workbooks.Open(self.folder + "\\" + file)
                    # wb.Application.Visible = False
                    wb.SaveAs(self.process_folder + file + "x", FileFormat=51)
                    # FileFormat = 51 is for .xlsx extension
                    # FileFormat = 56 is for .xls extension
                    print("%s has been saved as .xlsx" % file)
            wb.Close()
            excel.Application.Quit()
            print("All files have been saved as .xlsx")

    def new_columns(self):
        keyword = "Odau"
        file_list = os.listdir(self.process_folder)
        for file in file_list:
            if "log.txt" in file_list:
                print("列已添加")
                break
            else:
                if keyword in file:
                    print(file)
                    wb = openpyxl.load_workbook(self.process_folder + file)  # 打开文件
                    sheet_1 = wb.sheetnames[0]  # 获得第一个sheet名
                    data = pd.read_excel(self.process_folder + file, header=4)
                    print("新建肌电信号列")
                    data["Frame"] = data["Frame"] / 1000
                    data.rename(columns={"Frame": "time"}, inplace=True)
                    # print("Data:", data)
                    for i in range(1, 7):
                        EMG = ("EMG_" + str(i))
                        analog = ("Analog_" + str(i))
                        column = data[analog]
                        emg_mean = self.emg_mean(column)
                        data[EMG] = (data[analog] - emg_mean) / 2  # 除以500倍，×1000，mV
                    print("新建工作表RMS")

                    rms_data = pd.DataFrame()
                    for i in range(1, 7):
                        EMG_ = ("EMG_" + str(i))
                        RMS_ = ("RMS_" + str(i))
                        emg_rms_temp = self.RMS(data[EMG_])
                        # print("temp:", emg_rms_temp)
                        rms_data[RMS_] = emg_rms_temp

                    # 生成时间序列
                    frames = shape(rms_data)[0]
                    end_time = frames * self.win_set[0] / 1000
                    time = arange(0, end_time, self.win_set[0] / 1000)

                    df_rms = pd.DataFrame({"time": time,
                                           "RMS_1": rms_data['RMS_1'],
                                           "RMS_2": rms_data['RMS_2'],
                                           "RMS_3": rms_data['RMS_3'],
                                           "RMS_4": rms_data['RMS_4'],
                                           "RMS_5": rms_data['RMS_5'],
                                           "RMS_6": rms_data['RMS_6'],
                                           })
                    if self.upper_limb:
                        df_rms.rename(columns={"RMS_1": "RMS_1_三角肌前束",
                                               "RMS_2": "RMS_2_三角肌后束",
                                               "RMS_3": "RMS_3_肱二头肌",
                                               "RMS_4": "RMS_4_肱三头肌",
                                               "RMS_5": "RMS_5_桡侧腕屈肌",
                                               "RMS_6": "RMS_6_尺侧腕伸肌",
                                               }, inplace=True)
                    else:
                        df_rms.rename(columns={"RMS_1": "RMS_1_股直肌",
                                               "RMS_2": "RMS_2_股二头肌",
                                               "RMS_3": "RMS_3_半腱肌",
                                               "RMS_4": "RMS_4_股内侧肌",
                                               "RMS_5": "RMS_5_胫骨前肌",
                                               "RMS_6": "RMS_6_外侧腓肠肌",
                                               }, inplace=True)

                    # 分别写入两张表
                    writer = pd.ExcelWriter(self.process_folder + file)
                    data.to_excel(writer, sheet_name=sheet_1, index=False)
                    df_rms.to_excel(writer, sheet_name="RMS", index=False)
                    writer.save()

        self.log_create("log")

    def log_create(self, name):
        full_path = self.process_folder + name + '.txt'  # 也可以创建一个.doc的word文档
        file = open(full_path, 'w')
        file.write("新建列")

    def read_emg(self, path):
        return pd.read_excel(path, header=self.header)

    @staticmethod
    def emg_mean(column):
        return column.mean()

    @staticmethod
    def rect(raw):
        return np.fabs(raw)

    # 滑动时间窗
    def slid_window(self, raw):
        # 一次传入re_raw的一列
        length = len(raw)
        step = self.win_set[0]
        width = self.win_set[1]
        frame = int((length - width) / step + 1)  # 处理后的帧数
        sw = zeros((frame, width), dtype=float)  # 初始化矩阵，每行是一个时间窗
        i = 0
        while i < length - width + 1:
            # 每行新建空数组，作为window
            window = array([])
            for j in range(0, width):
                # 每列从i添加到i+width
                window = append(window, raw[i + j])
            # 生成sw矩阵
            sw[int(i / step)] = window
            i += step
        return sw

    def RMS(self, raw):
        sw = self.slid_window(raw)
        frame, width = shape(sw)
        RMS = zeros(frame, dtype=float)
        for j in range(0, frame):
            rms_p = 0
            for k in range(0, width):
                rms_p = rms_p + sw[j][k] ** 2  # 平方和
            RMS[j] = (rms_p / width) ** 0.5  # 开方
        return RMS

    def AEMG(self, re_raw):
        pass

    def rms_plot(self):
        keyword = "Odau"
        file_list = os.listdir(self.process_folder)

        font = FontProperties(fname=r"c:\windows\fonts\simsun.ttc", size=15)
        for file in file_list:
            if "plot.txt" in file_list:
                print("列已添加")
                break
            else:
                if keyword in file:
                    rms = pd.read_excel(self.process_folder + file, sheet_name="RMS")
                    rms.plot(x="time", y=[1, 2, 3, 4, 5, 6])
                    plt.title(file[-11:-5], font=font)
                    plt.ylabel('RMS/mV')
                    plt.show()
        print("图已添加")

    def plot_in_excel(self):
        keyword = "Odau"
        file_list = os.listdir(self.process_folder)
        for file in file_list:
            if keyword in file:
                wb = openpyxl.load_workbook(self.folder + "\\Process\\" + file)
                print("open:", self.folder + "\\Process\\" + file)
                print(wb.sheetnames)

                # ws = wb.get_sheet_by_name("RMS")
                ws = wb["RMS"]
                max_row = ws.max_row
                print("max_row", max_row)

                chart1 = LineChart()
                # chart1.type = "col"
                chart1.style = 10
                chart1.title = "RMS"

                labels = Reference(ws, min_col=2, max_col=7, min_row=1)
                cats = Reference(ws, min_col=1, min_row=2, max_row=max_row + 1)
                data1 = Reference(ws, min_col=2, min_row=1, max_row=max_row + 1)
                data2 = Reference(ws, min_col=3, min_row=1, max_row=max_row + 1)
                data3 = Reference(ws, min_col=4, min_row=1, max_row=max_row + 1)
                data4 = Reference(ws, min_col=5, min_row=1, max_row=max_row + 1)
                data5 = Reference(ws, min_col=6, min_row=1, max_row=max_row + 1)
                data6 = Reference(ws, min_col=7, min_row=1, max_row=max_row + 1)

                chart1.add_data(data1, titles_from_data=True)
                chart1.add_data(data2, titles_from_data=True)
                chart1.add_data(data3, titles_from_data=True)
                chart1.add_data(data4, titles_from_data=True)
                chart1.add_data(data5, titles_from_data=True)
                chart1.add_data(data6, titles_from_data=True)

                chart1.set_categories(cats)
                # s1 = chart1.series[0:6]
                # s1.graphicalProperties.line.width = 1000
                chart1.width = 40
                chart1.height = 20
                chart1.x_axis.title = "time/s"
                chart1.y_axis.title = "RMS/mV"
                chart1.y_axis.scaling.max = 0.2


                ws.add_chart(chart1, "I5")
                wb.save(self.folder + "\\Process\\" + file)

    def run(self):
        # 1.另存为高版本，测试文件重命名对应动作
        self.save_as_high_ver()
        self.rename_test_file()
        # # 2.肌电信号预处理，矫正零偏，计算RMS
        self.new_columns()
        # # # 3. 插入图像
        self.plot_in_excel()
        # self.rms_plot()


def main():
    # folder = r"D:\code\运动能力分析实验\0921wrj_2020_09_21_140406"
    folder = r"D:\code\运动能力分析实验\0921wrj_2020_09_21_144822"
    subject = r"D:\code\运动能力分析实验\邬如靖.xlsx"

    EP = EMGProcess(folder, subject)
    EP.run()


if __name__ == '__main__':
    main()
