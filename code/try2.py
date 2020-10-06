# author:wzt
import win32com.client as win32

from numpy import *
import openpyxl
import pandas as pd
# wb = openpyxl.Workbook()
# sheet = wb.active
# print(sheet.title)
# sheet.title = "wzt"
# wb.create_sheet(title="RMS")
# print(wb.sheetnames)
# wb.save("wzt.xlsx")


df1 = pd.DataFrame({'a':[1,2,3], 'b':[4,5,6]})
df2 = pd.DataFrame({'a':[7,8,9], 'b':[6,5,4]}, columns=['a', 'b'])
print(df1)
writer = pd.ExcelWriter('a.xls')
df1.to_excel(writer, sheet_name='sheet1', index=False)
df2.to_excel(writer, sheet_name='sheet2', index=False)
writer.save()


# df1 = pd.DataFrame({'a':[1,2,3], 'b':[4,5,6]})
df2 = pd.DataFrame({'a':[7,8,9], 'b':[6,5,4]}, columns=['a', 'b'])
# print(df1)
writer = pd.ExcelWriter('a.xls')
# df1.to_excel(writer, sheet_name='sheet1', index=False)
df2.to_excel(writer, sheet_name='sheet2', index=False)
writer.save()



